#!/usr/bin/env python3
"""
Script to create a simple test experiment template.
"""

import openpyxl
from openpyxl import Workbook

def create_simple_test():
    """Create a minimal test experiment template."""
    wb = Workbook()

    # Sheet: settings
    ws_settings = wb.active
    ws_settings.title = 'settings'
    ws_settings.append(['settings_label', 'value'])
    ws_settings.append(['session_id', 'simple_test'])
    ws_settings.append(['model_info', 'gpt-4o'])  # Use standard model name
    ws_settings.append(['hf_inference_endpoint', ''])  # Required field
    ws_settings.append(['temperature', 0])
    ws_settings.append(['num_subjects_per_group', 1])
    ws_settings.append(['num_groups', 1])
    ws_settings.append(['max_num_rounds', 2])  # Need at least 2 rounds for the question
    ws_settings.append(['treatment_assignment_strategy', 'manual'])
    ws_settings.append(['treatment_column', 'treatment'])  # Required field
    ws_settings.append(['group_assignment_strategy', 'manual'])  # Required field
    ws_settings.append(['group_column', 'group'])  # Required field
    ws_settings.append(['role_assignment_strategy', 'manual'])  # Required field
    ws_settings.append(['role_column', 'role'])  # Required field
    ws_settings.append(['random_seed', 42])  # Required field
    ws_settings.append(['build_profile_qna', False])  # Required field
    ws_settings.append(['build_profile_backstories', False])  # Required field

    # Sheet: treatment
    ws_treatment = wb.create_sheet('treatment')
    ws_treatment.append(['treatment_label', 'value'])
    ws_treatment.append(['control', 'control'])

    # Sheet: role
    ws_role = wb.create_sheet('role')
    ws_role.append(['role_label', 'value'])
    ws_role.append(['participant', 'You are a helpful assistant. Answer the question: What is 2+2?'])

    # Sheet: prompt (complex format with many required columns)
    # Need both context and question types to trigger API calls
    ws_prompt = wb.create_sheet('prompt')
    ws_prompt.append([
        'round_id', 'type', 'round_order', 'is_adapted', 'human_text', 'llm_text',
        'response_name', 'response_type', 'response_options', 'randomize_response_order',
        'validate_response', 'generate_speculation_score', 'format_response'
    ])
    # Context prompt
    ws_prompt.append([
        0, 'context', 0, False,
        'Simple Test\nYou will be asked a simple math question.',
        'Simple Test\nYou will be asked a simple math question.',
        'experiment_context', 'context', None, False, False, False, False
    ])
    # Question prompt that requires an answer (must be private_question type)
    ws_prompt.append([
        1, 'private_question', 1, False,
        'What is 2+2?',
        'What is 2+2?',
        'answer', 'text', None, False, False, False, False
    ])

    # Sheet: profile (needs at least one data row)
    ws_profile = wb.create_sheet('profile')
    # Profile sheet has many columns - using minimal set
    ws_profile.append(['ID', 'treatment', 'group', 'role'])
    ws_profile.append(['1', 'control', 1, 'participant'])

    # Sheet: constant (empty but required)
    ws_constant = wb.create_sheet('constant')
    ws_constant.append(['constant_label', 'value'])

    wb.save('simple_test.xlsx')
    print('✅ Created simple_test.xlsx')
    print('   This is a minimal test experiment with:')
    print('   - 1 group, 1 participant')
    print('   - 1 round')
    print('   - Simple question: What is 2+2?')
    print('   - Model: gpt-4o-mini')
    print('\n   Run it with:')
    print('   poetry run python run_experiment.py simple_test.xlsx test')

if __name__ == "__main__":
    create_simple_test()

